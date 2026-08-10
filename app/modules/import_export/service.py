from __future__ import annotations

import json
import re
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.errors import BusinessError
from app.models import (
    AuditLog,
    Job,
    KnowledgeObject,
)
from app.models.base import utc_now
from app.modules.knowledge.service import _create_revision
from app.modules.relation.service import _create_edge
from app.schemas.catalog import GRADE_TERM_LABELS, KnowledgeScope, KnowledgeType

IMPORT_HEADERS = (
    "canonical_id",
    "type",
    "name",
    "grade_term",
    "scope",
    "pep24_path",
    "ocr_signals",
    "exercise_signature",
    "prerequisites",
)
GRADE_TERM_BY_LABEL = {value: key for key, value in GRADE_TERM_LABELS.items()}
KNOWLEDGE_TYPES = set(KnowledgeType.__args__)
SCOPES = set(KnowledgeScope.__args__)
CANONICAL_ID = re.compile(r"^1\d{7}$")
MAX_IMPORT_BYTES = 10 * 1024 * 1024
MAX_IMPORT_ROWS = 1000


def _error(row: int, field: str, reason: str, suggestion: str) -> dict[str, Any]:
    return {
        "rowNumber": row,
        "field": field,
        "reason": reason,
        "suggestion": suggestion,
    }


def _json_array(value: str, row: int, field: str, errors: list[dict[str, Any]]):
    if not value:
        return []
    try:
        result = json.loads(value)
    except json.JSONDecodeError:
        result = None
    if not isinstance(result, list):
        errors.append(_error(row, field, "必须是 JSON 数组", "例如填写 []"))
        return []
    if any(not isinstance(item, str) for item in result):
        errors.append(_error(row, field, "数组元素必须是字符串", "请检查 JSON 数组"))
        return []
    return list(dict.fromkeys(item.strip() for item in result))


def _validate(content: bytes):
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise BusinessError("FILE_INVALID", "Excel 文件无法读取", 422) from exc
    if "knowledge_points" not in workbook.sheetnames:
        raise BusinessError("VALIDATION_FAILED", "缺少 knowledge_points 工作表", 422)
    sheet = workbook["knowledge_points"]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise BusinessError("VALIDATION_FAILED", "Excel 没有数据", 422)
    headers = tuple(
        str(value).strip() if value is not None else "" for value in rows[0]
    )
    if headers != IMPORT_HEADERS:
        raise BusinessError(
            "VALIDATION_FAILED",
            "Excel 表头不匹配",
            422,
            {"expected": list(IMPORT_HEADERS), "actual": list(headers)},
        )
    errors: list[dict[str, Any]] = []
    records = []
    seen: set[str] = set()
    for row_number, row in enumerate(rows[1:], 2):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        value = {
            field: str(row[index]).strip()
            if index < len(row) and row[index] is not None
            else ""
            for index, field in enumerate(IMPORT_HEADERS)
        }
        for field in ("canonical_id", "type", "name", "grade_term", "scope"):
            if not value[field]:
                errors.append(_error(row_number, field, "不能为空", "请补充字段"))
        canonical_id = value["canonical_id"]
        if canonical_id and not CANONICAL_ID.fullmatch(canonical_id):
            errors.append(
                _error(
                    row_number,
                    "canonical_id",
                    "格式无效",
                    "使用以 1 开头的 8 位纯数字 ID",
                )
            )
        if canonical_id in seen:
            errors.append(
                _error(row_number, "canonical_id", "文件内重复", "ID 只能出现一次")
            )
        seen.add(canonical_id)
        if value["type"] and value["type"] not in KNOWLEDGE_TYPES:
            errors.append(
                _error(row_number, "type", "枚举值无效", "请使用系统类型 key")
            )
        grade_term = GRADE_TERM_BY_LABEL.get(value["grade_term"], value["grade_term"])
        if grade_term not in GRADE_TERM_LABELS:
            errors.append(
                _error(row_number, "grade_term", "年级/学期无效", "请使用教材年级")
            )
        if value["scope"] and value["scope"] not in SCOPES:
            errors.append(
                _error(row_number, "scope", "枚举值无效", "请使用 core 或 supplement")
            )
        ocr_signals = _json_array(
            value["ocr_signals"], row_number, "ocr_signals", errors
        )
        prerequisites = _json_array(
            value["prerequisites"], row_number, "prerequisites", errors
        )
        if canonical_id in prerequisites:
            errors.append(
                _error(row_number, "prerequisites", "不能引用自身", "删除自身知识点 ID")
            )
        records.append(
            {
                **value,
                "grade_term_code": grade_term,
                "ocr_signals": ocr_signals,
                "prerequisites": prerequisites,
                "row_number": row_number,
            }
        )
    return len(records), errors, records


def _job_response(job: Job) -> dict[str, Any]:
    payload = dict(job.payload_json or {})
    errors = payload.get("errors", [])
    return {
        "jobId": str(job.id),
        "jobType": job.job_type,
        "status": job.status,
        "totalCount": job.total_count,
        "successCount": job.success_count,
        "failureCount": job.failure_count,
        "progressPercent": 100 if job.status in {"success", "failed"} else 0,
        "errorFileId": job.error_file,
        "canRetry": False,
        "errorCount": len(errors),
        "committed": bool(payload.get("committed")),
        "canCommit": job.status == "success" and not payload.get("committed"),
    }


def job_response(job: Job) -> dict[str, Any]:
    return _job_response(job)


def create_import_job(
    session: Session, filename: str, content: bytes, actor: str, request_id: str
) -> Job:
    if not filename.lower().endswith(".xlsx"):
        raise BusinessError("FILE_INVALID", "仅支持 .xlsx 文件", 422)
    if len(content) > MAX_IMPORT_BYTES:
        raise BusinessError("FILE_TOO_LARGE", "Excel 文件不能超过 10 MB", 413)
    total, errors, records = _validate(content)
    if total > MAX_IMPORT_ROWS:
        raise BusinessError("VALIDATION_FAILED", "单批导入不能超过 1000 行", 422)
    job = Job(
        job_type="import",
        status="failed" if errors else "success",
        total_count=total,
        success_count=0 if errors else total,
        failure_count=total if errors else 0,
        payload_json={"filename": filename, "errors": errors, "rows": records},
        created_by=actor,
    )
    session.add(job)
    session.flush()
    session.add(
        AuditLog(
            actor_id=actor,
            action="import.create",
            entity_type="job",
            entity_key=str(job.id),
            affected_knowledge_base_ids=[],
            request_id=request_id,
            created_at=utc_now(),
        )
    )
    session.commit()
    return job


def get_job(session: Session, job_id: int) -> Job:
    job = session.get(Job, job_id)
    if job is None:
        raise BusinessError("NOT_FOUND", "任务不存在", 404)
    return job


def list_job_errors(
    session: Session, job_id: int, page_num: int, page_size: int
) -> tuple[int, list[dict[str, Any]]]:
    errors = (get_job(session, job_id).payload_json or {}).get("errors", [])
    start = (page_num - 1) * page_size
    return len(errors), errors[start : start + page_size]


def commit_import_job(
    session: Session, job_id: int, actor: str, request_id: str
) -> Job:
    job = get_job(session, job_id)
    payload = dict(job.payload_json or {})
    if payload.get("committed"):
        return job
    if job.status != "success":
        raise BusinessError("VALIDATION_FAILED", "预校验未通过，不能提交", 422)
    records = payload.get("rows", [])
    ids = [row["canonical_id"] for row in records]
    existing = set(
        session.scalars(
            select(KnowledgeObject.canonical_id).where(
                KnowledgeObject.canonical_id.in_(ids)
            )
        )
    )
    if existing:
        raise BusinessError(
            "CONFLICT", f"知识点已存在：{', '.join(sorted(existing))}", 409
        )
    all_ids = set(ids)
    referenced = {item for row in records for item in row["prerequisites"]}
    missing = (
        referenced
        - all_ids
        - set(
            session.scalars(
                select(KnowledgeObject.canonical_id).where(
                    KnowledgeObject.canonical_id.in_(referenced - all_ids)
                )
            )
        )
    )
    if missing:
        raise BusinessError(
            "VALIDATION_FAILED",
            f"前置关系引用不存在：{', '.join(sorted(missing))}",
            422,
        )
    by_canonical: dict[str, KnowledgeObject] = {}
    try:
        for row in records:
            knowledge = KnowledgeObject(
                canonical_id=row["canonical_id"],
                created_by=actor,
                updated_by=actor,
            )
            session.add(knowledge)
            session.flush()
            _create_revision(
                session,
                knowledge,
                {
                    "name": row["name"],
                    "type": row["type"],
                    "grade_term": row["grade_term_code"],
                    "scope": row["scope"],
                    "ocr_signals": row["ocr_signals"],
                    "exercise_signature": row["exercise_signature"] or None,
                },
                actor,
            )
            by_canonical[knowledge.canonical_id] = knowledge
        for row in records:
            target = by_canonical[row["canonical_id"]]
            for prerequisite_id in row["prerequisites"]:
                source = by_canonical.get(prerequisite_id) or session.scalar(
                    select(KnowledgeObject).where(
                        KnowledgeObject.canonical_id == prerequisite_id
                    )
                )
                _create_edge(session, source, target, "prerequisite", None, actor)
        payload["committed"] = True
        job.payload_json = payload
        session.add(
            AuditLog(
                actor_id=actor,
                action="import.commit",
                entity_type="job",
                entity_key=str(job.id),
                affected_knowledge_base_ids=[],
                request_id=request_id,
                created_at=utc_now(),
            )
        )
        session.commit()
    except Exception:
        session.rollback()
        raise
    return job
